# -*- coding: utf-8 -*-
"""
개별종목 보유 ETF 분석기  (병렬 크롤링 버전)
============================================================
국내 상장 ETF 전체의 PDF(구성종목)를 받아, 각 개별 상장주식이
ETF들에 의해 얼마나 보유되고 있는지를 집계한다.

산출물(엑셀):
  1_시총대비편입비중 : 시총 대비 ETF 편입 비중이 큰 종목
  2_편입급증         : 비교 기준일 대비 편입 비중이 늘어난 종목
  3_조회             : 종목코드/종목명을 입력하면 편입비중이 자동으로 뜸
  4_전체데이터       : 전 종목 집계 결과(조회 시트의 데이터 원본)

방법론
------
- 모든 ETF PDF에서 "실제 상장 개별주식"인 구성종목만 추림
  (채권/선물/현금/타ETF는 시총 유니버스에 없어 자동 제외, 혼합형 개별종목은 포함).
- 종목별 ETF 보유액 = Σ( ETF 순자산(NAV×상장좌수) × 구성비중 )
- 편입비중(%) = 종목별 ETF 보유액 합 / 시가총액 × 100

사용법: README.md 참고. (KRX 로그인 필요 - krx_계정.txt)
"""

import os
import sys
import json
import time
import socket as _socket
_socket.setdefaulttimeout(30)  # 요청 1건 무한대기(hang) 방지
import warnings
import datetime as dt
from concurrent.futures import ThreadPoolExecutor, as_completed

warnings.filterwarnings("ignore")  # pykrx 내부의 무해한 pandas 경고 숨김
AUTO = "--auto" in sys.argv  # 무인 실행(스케줄러): 날짜 안 묻고 기본값, 끝에 멈추지 않음

# --year 2026 / --year=2026 : 그 해 거래일 전체를 이어받기(배치) 모드로 수집
YEAR = None
for _i, _a in enumerate(sys.argv):
    if _a == "--year" and _i + 1 < len(sys.argv):
        YEAR = sys.argv[_i + 1]
    elif _a.startswith("--year="):
        YEAR = _a.split("=", 1)[1]
INCLUDE_ETF = "--withetf" in sys.argv  # ETF가 보유한 '다른 ETF'도 종목처럼 포함

import pandas as pd
from tqdm import tqdm

# ── KRX 로그인 (2025-12-27 개편 이후 KRX 데이터는 로그인 필수) ──────────
_ACCOUNT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "krx_계정.txt")
if not (os.environ.get("KRX_ID") and os.environ.get("KRX_PW")):
    if os.path.exists(_ACCOUNT_FILE):
        _lines = [ln.strip() for ln in open(_ACCOUNT_FILE, encoding="utf-8")
                  if ln.strip() and not ln.strip().startswith("#")]
        if len(_lines) >= 2 and "여기에" not in _lines[0]:
            os.environ["KRX_ID"] = _lines[0]
            os.environ["KRX_PW"] = _lines[1]
KRX_LOGGED_IN = bool(os.environ.get("KRX_ID") and os.environ.get("KRX_PW"))

# ── (중요) pykrx import 시 로그인하며 세션이 만들어지므로, 그 전에 풀을 키운다 ──
# HTTPAdapter 기본 풀 크기를 WORKERS(=16)에 맞춰 둔다. 이후 만들어지는 모든
# requests.Session()이 자동으로 큰 풀을 갖게 되어 'pool is full(size:10)' 경고가 사라짐.
try:
    from requests.adapters import HTTPAdapter as _HTTPAdapter_early
    _POOL_EARLY = 10   # WORKERS와 동일하게 유지
    _orig_ad_init_early = _HTTPAdapter_early.__init__
    def _big_ad_init_early(self, *a, **kw):
        kw.setdefault("pool_connections", _POOL_EARLY)
        kw.setdefault("pool_maxsize", _POOL_EARLY)
        return _orig_ad_init_early(self, *a, **kw)
    _HTTPAdapter_early.__init__ = _big_ad_init_early
except Exception:
    pass

try:
    from pykrx import stock
    from pykrx.website.krx.etx.core import 전종목시세_ETF
except Exception as _imp_e:
    # KRX가 로그인 응답을 HTML(차단페이지)로 주면 여기서 터진다 = throttle.
    # 종료코드 3으로 나가면 bat이 40분 쉬었다 재시도한다(무한 15초 재시도 방지).
    print("KRX 초기화/로그인 실패(throttle 의심) - 40분 쉬었다 재시도:", _imp_e)
    sys.exit(3)

# ============================================================
# CONFIG  (여기만 바꾸면 됨)
# ============================================================

BASE_DATE = ""              # 분석 기준일. "" 이면 최근 영업일 자동. "YYYYMMDD"
COMPARE_DATE = "20251230"   # 비교 기준일(2025년 마지막 거래일)

# 동시요청: IP 차단 이후 안전하게 10 고정(요청 속도를 낮춰 throttle 회피).
WORKERS = 10
RETRIES = 4        # PDF 조회 실패/빈응답 시 재시도 횟수
TOP_N = 300        # 시트1·2 상위 개수 (None 이면 전체)
SLEEP = 0.1        # 각 요청 후 추가 대기(초). throttle 완화용.
ASK_DATES = True   # (현재 미사용) 호환용
REST_BETWEEN = 30  # 날짜 사이 쉬는 시간(초). throttle 누적 방지.
BATCH_PER_RUN = 1      # 한 번 실행에 1일치만(매 날짜 새 로그인 → 세션 만료 회피)
FAIL_RATIO_LIMIT = 0.05  # PDF 실패율이 이보다 높으면 throttle로 간주 → 저장 안 하고 다음에 재시도
MAKE_EXCEL = False  # 엑셀 생성 안 함(웹 데이터만). True면 수동 실행 시 엑셀도 생성.
WRITE_CACHE = False # 집계 캐시(.pkl) 저장 안 함. 이어받기는 웹파일 기준이라 불필요+디스크 낭비 방지.

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(OUTPUT_DIR, "cache")
os.makedirs(CACHE_DIR, exist_ok=True)
WEB_DIR = os.path.join(OUTPUT_DIR, "data")   # 웹용 json 출력 폴더
os.makedirs(WEB_DIR, exist_ok=True)

# ── pykrx 연결풀을 WORKERS 크기로 확장 (기본 10 제한 해제) ──
# pykrx는 로그인(refresh)마다 requests.Session()을 새로 만들어 풀이 10으로
# 초기화된다. refresh 직후 큰 풀 어댑터를 매번 다시 끼워 16 동시연결을 보장.
try:
    from requests.adapters import HTTPAdapter as _HTTPAdapter
    from pykrx.website.comm import auth as _auth
    _POOL = 10

    def _mount_big_pool(_sess):
        _ad = _HTTPAdapter(pool_connections=_POOL, pool_maxsize=_POOL, max_retries=0)
        _sess.mount("https://", _ad)
        _sess.mount("http://", _ad)

    _orig_refresh = _auth.KRXSession.refresh

    def _patched_refresh(self, login_id, login_pw):
        try:
            return _orig_refresh(self, login_id, login_pw)
        finally:
            try:
                _mount_big_pool(self.session)
            except Exception:
                pass

    _auth.KRXSession.refresh = _patched_refresh
except Exception:
    pass


# ============================================================
# 유틸 / 로그인
# ============================================================

def _to_num(s):
    if isinstance(s, str):
        s = s.replace(",", "").strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def resolve_base_date():
    if BASE_DATE:
        return BASE_DATE
    try:
        from pykrx.stock import get_nearest_business_day_in_a_week
        return get_nearest_business_day_in_a_week()
    except Exception:
        return dt.date.today().strftime("%Y%m%d")


def _valid_date(s):
    s = s.strip()
    return s.isdigit() and len(s) == 8


def ask_dates():
    """실행 시 기준일/비교일을 입력받는다. 엔터=기본값."""
    base_default = BASE_DATE if BASE_DATE else resolve_base_date()
    cmp_default = COMPARE_DATE
    if not ASK_DATES or AUTO:
        return base_default, cmp_default
    print("\n날짜를 입력하세요 (YYYYMMDD, 그냥 엔터치면 기본값 사용)")
    try:
        b = input(f"  기준일(분석 시점) [엔터={base_default}]: ").strip()
        c = input(f"  비교 기준일       [엔터={cmp_default}]: ").strip()
    except EOFError:
        return base_default, cmp_default
    base = b if _valid_date(b) else base_default
    cmp_ = c if _valid_date(c) else cmp_default
    if b and not _valid_date(b):
        print(f"  (기준일 형식이 이상해서 {base} 로 진행)")
    if c and not _valid_date(c):
        print(f"  (비교 기준일 형식이 이상해서 {cmp_} 로 진행)")
    return base, cmp_


def ask_dates_multi():
    """크롤링할 날짜들을 여러 개 입력받는다. 엔터=최근 영업일 1개."""
    import re
    print("\n크롤링할 날짜를 입력하세요. 여러 개면 띄어쓰기/콤마로 구분.")
    print("  예) 20260430 20260529 20260625    (그냥 엔터 = 최근 영업일)")
    try:
        raw = input("날짜들: ").strip()
    except EOFError:
        raw = ""
    if not raw:
        return [resolve_base_date()]
    toks = [t for t in re.split(r"[\s,]+", raw) if t]
    good = [t for t in toks if _valid_date(t)]
    bad = [t for t in toks if not _valid_date(t)]
    if bad:
        print("  (형식이 이상해 무시:", bad, ")")
    return good or [resolve_base_date()]


def ensure_krx_login():
    if KRX_LOGGED_IN:
        return True
    if not os.path.exists(_ACCOUNT_FILE):
        with open(_ACCOUNT_FILE, "w", encoding="utf-8") as f:
            f.write("# KRX 로그인 정보 파일\n")
            f.write("# data.krx.co.kr 에서 무료 회원가입 후, 아래 두 줄을\n")
            f.write("# 본인 아이디/비밀번호로 바꿔 저장하세요. (# 줄은 무시)\n")
            f.write("여기에_KRX_아이디\n")
            f.write("여기에_KRX_비밀번호\n")
    print("\n" + "=" * 56)
    print(" KRX 로그인 정보가 없습니다.")
    print(" 1) data.krx.co.kr 에서 (무료) 회원가입")
    print(" 2) 같은 폴더의 'krx_계정.txt' 1줄=아이디, 2줄=비밀번호 저장")
    print(" 3) 다시 실행")
    print("=" * 56)
    return False


# ============================================================
# 데이터 수집
# ============================================================

def get_market_cap(date):
    """시가총액 조회. KRX가 빈 응답/오류를 주면 빈 DataFrame 반환(죽지 않음). 재시도 포함."""
    for _ in range(3):
        try:
            df = stock.get_market_cap_by_ticker(date, market="ALL", alternative=True)
            if df is not None and not df.empty and "시가총액" in df.columns:
                return df[df["시가총액"] > 0]
        except Exception:
            pass
        time.sleep(1.0)
    return pd.DataFrame(columns=["시가총액"])


def get_etf_meta(date):
    """ETF별 (순자산 dict, 이름 dict)을 한 번의 호출로 반환."""
    raw = 전종목시세_ETF().fetch(date)
    sizes, names = {}, {}
    if raw is None or raw.empty:
        return sizes, names
    for _, r in raw.iterrows():
        tk = str(r.get("ISU_SRT_CD", "")).zfill(6)
        amt = _to_num(r.get("NETASST_TOTAMT")) or _to_num(r.get("NAV_TOTAMT"))
        aum = amt if amt > 0 else _to_num(r.get("NAV")) * _to_num(r.get("LIST_SHRS"))
        if aum <= 0:
            aum = _to_num(r.get("MKTCAP"))
        sizes[tk] = aum
        nm = r.get("ISU_ABBRV", tk)
        names[tk] = nm if isinstance(nm, str) and nm else tk
    return sizes, names


def get_etf_sizes(date):
    """ETF별 순자산(원). (종목검증.py 호환)"""
    return get_etf_meta(date)[0]


def get_etf_universe(date):
    """ETF를 '종목'처럼 쓰기 위한 시총/종가/상장좌수 (ETF-in-ETF 보유 반영용).
    {코드: {"종가","시가총액","상장주식수","name"}} 반환."""
    try:
        raw = 전종목시세_ETF().fetch(date)
    except Exception:
        return None
    if raw is None or raw.empty:
        return None
    out = {}
    for _, r in raw.iterrows():
        c = str(r.get("ISU_SRT_CD", "")).zfill(6)
        if not c or c == "000000":
            continue
        mkt = _to_num(r.get("MKTCAP"))
        if mkt <= 0:
            mkt = _to_num(r.get("NAV")) * _to_num(r.get("LIST_SHRS"))
        prc = _to_num(r.get("TDD_CLSPRC")) or _to_num(r.get("NAV"))
        shr = _to_num(r.get("LIST_SHRS"))
        nm = r.get("ISU_ABBRV", c)
        out[c] = {"종가": prc, "시가총액": mkt, "상장주식수": shr,
                  "name": nm if isinstance(nm, str) and nm else c}
    return out


def get_etf_list_robust(date, tries=4):
    """ETF 목록을 여러 번 받아 '가장 많이 받힌' 결과 사용(부분 수신 방지)."""
    best = []
    for _ in range(tries):
        try:
            lst = stock.get_etf_ticker_list(date)
        except Exception:
            lst = []
        if lst and len(lst) > len(best):
            best = lst
        time.sleep(0.4)
    return best


def _fetch_pdf(etf, date):
    """PDF 1건 조회. 빈 응답/오류 모두 재시도(세션 만료 대비). (etf, DataFrame|None)."""
    for i in range(RETRIES + 1):
        try:
            pdf = stock.get_etf_portfolio_deposit_file(etf, date)
            if pdf is not None and not pdf.empty and "비중" in pdf.columns:
                if SLEEP:
                    time.sleep(SLEEP)
                return etf, pdf
            # 빈 응답: 합성형일 수도, 세션만료로 인한 일시 빈응답일 수도 →
            # 마지막 시도까지 비어있으면 그때 None 처리(=진짜 비어있음으로 간주)
        except Exception:
            pass
        time.sleep(0.4 * (i + 1))
    return etf, None


class ThrottledError(Exception):
    """PDF 실패율이 너무 높음(=KRX throttle). 저장하지 않고 다음에 재시도."""
    pass


def aggregate_for_date(date, mcap_index):
    """종목별 ETF 보유액 병렬 집계 + ETF별 상세. 캐시: cache/agg_v6_{date}.pkl
    반환: (agg_df, detail_df). 실패율 높으면 ThrottledError."""
    cache_path = os.path.join(CACHE_DIR, f"agg_v6_{date}.pkl")
    if os.path.exists(cache_path):
        print(f"  [cache] {date} 집계 캐시 사용")
        d = pd.read_pickle(cache_path)
        return d["agg"], d["detail"]

    print(f"  ETF 목록/순자산 조회중... ({date})")
    etf_list = get_etf_list_robust(date)
    sizes, etf_names = get_etf_meta(date)
    targets = [e for e in etf_list if sizes.get(e, 0.0) > 0]
    print(f"  ETF {len(etf_list)}개 (보유액>0 {len(targets)}개). "
          f"PDF 병렬 크롤링 시작 (worker={WORKERS}).")

    hold_value, hold_count = {}, {}
    detail_rows = []
    fail = 0
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = {ex.submit(_fetch_pdf, e, date): e for e in targets}
        for fut in tqdm(as_completed(futures), total=len(futures),
                        desc=f"  PDF {date}", unit="etf"):
            etf, pdf = fut.result()
            if pdf is None:
                fail += 1
                continue
            aum = sizes[etf]
            enm = etf_names.get(etf, etf)
            # --- 1CU 순자산총액(분모) 파악 ---
            # 해외혼합 ETF: '설정현금액' 행의 시가총액 = 1CU 순자산총액(해외분 포함).
            # 국내전용 ETF: 설정현금액 행이 없음 → 시가총액 칸 전체 합 = 순자산.
            setjeong = 0.0   # 설정현금액(= 총 NAV)
            cash_amt = 0.0   # 원화현금
            sigtot = 0.0     # 시가총액 칸 전체 합(설정현금액 제외)
            for _, row in pdf.iterrows():
                nm = str(row.get("구성종목명", "")).strip()
                sga = _to_num(row.get("시가총액"))
                if nm == "설정현금액":
                    setjeong = sga
                else:
                    sigtot += sga
                    if nm == "원화현금":
                        cash_amt = sga
            denom = setjeong if setjeong > 0 else sigtot

            for tk, row in pdf.iterrows():
                nm = str(row.get("구성종목명", "")).strip()
                if nm == "설정현금액":
                    continue   # 분모 전용 — 보유 항목 아님
                if nm == "원화현금":   # 현금: 종목 아님 → 종목집계 제외, 현금비중만 보존
                    if denom > 0 and cash_amt > 0:
                        cw = cash_amt / denom
                        detail_rows.append(("원화현금", etf, enm, cw * 100.0, aum * cw, aum, 0.0))
                    continue
                tk = str(tk).zfill(6)
                if tk not in mcap_index:
                    continue
                bijung = _to_num(row.get("비중"))
                sga = _to_num(row.get("시가총액"))
                gye = _to_num(row.get("계약수"))
                if bijung > 0:
                    w = bijung / 100.0                 # KRX 공식 비중(패시브 등)
                elif denom > 0 and sga > 0:
                    w = sga / denom                    # 해외혼합: 시가총액 ÷ 설정현금액
                else:
                    w = 0.0
                val = aum * w
                hold_value[tk] = hold_value.get(tk, 0.0) + val
                hold_count[tk] = hold_count.get(tk, 0) + 1
                detail_rows.append((tk, etf, enm, w * 100.0, val, aum, gye))

    if fail:
        print(f"  (PDF 비어있음/실패 {fail}건 - 합성·현금형 등, 무시)")

    # 실패율 가드: 너무 많이 실패했으면 throttle로 보고 '저장 안 함' → 다음에 재시도
    fail_ratio = fail / max(1, len(targets))
    if fail_ratio > FAIL_RATIO_LIMIT:
        print(f"  ⚠️ 실패율 {fail_ratio:.0%} ({fail}/{len(targets)}) — throttle로 판단. "
              f"이 날짜는 저장하지 않고 다음에 다시 받습니다.")
        raise ThrottledError(date)

    agg = pd.DataFrame({
        "etf보유액": pd.Series(hold_value, dtype="float64"),
        "보유etf수": pd.Series(hold_count, dtype="float64"),
    }).fillna(0)
    agg.index.name = "티커"
    detail = pd.DataFrame(detail_rows,
                          columns=["종목코드", "ETF코드", "ETF명", "비중", "보유액", "AUM", "계약수"])
    if WRITE_CACHE:
        pd.to_pickle({"agg": agg, "detail": detail}, cache_path)
    return agg, detail


def build_name_map(tickers, known=None):
    cache_path = os.path.join(CACHE_DIR, "names.json")
    names = {}
    if os.path.exists(cache_path):
        try:
            with open(cache_path, encoding="utf-8") as f:
                names = json.load(f)
        except Exception:
            names = {}   # 캐시가 깨졌으면 새로 시작
    # 문자열이 아닌 잘못된 값은 버림(과거 버그로 섞였을 수 있음)
    names = {k: v for k, v in names.items() if isinstance(v, str)}
    # 미리 아는 이름(ETF 등)은 조회 없이 채움 → pykrx의 ETF 종목명 조회 스팸 방지
    if known:
        names.update({k: v for k, v in known.items() if isinstance(v, str) and v})
    missing = [t for t in tickers if t not in names]

    def _nm(t):
        try:
            nm = stock.get_market_ticker_name(t)
            if isinstance(nm, str) and nm:
                return t, nm
        except Exception:
            pass
        return t, t   # ETF 등 종목명이 안 나오면 코드로 대체(뒤에서 ETF명 보강)

    if missing:
        with ThreadPoolExecutor(max_workers=WORKERS) as ex:
            for t, nm in tqdm(ex.map(_nm, missing), total=len(missing),
                              desc="  종목명", unit="stk"):
                names[t] = nm
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(names, f, ensure_ascii=False)
    return names


# ============================================================
# 분석 결합
# ============================================================

def make_ratio_table(agg, mcap, names):
    joined = agg.join(mcap[["시가총액"]], how="inner")
    tickers = list(joined.index)
    return pd.DataFrame({
        "티커": tickers,
        "종목명": [names.get(t, t) for t in tickers],
        "시가총액": joined["시가총액"].to_numpy(),
        "etf보유액": joined["etf보유액"].to_numpy(),
        "편입비중(%)": (joined["etf보유액"] / joined["시가총액"] * 100).to_numpy(),
        "보유etf수": joined["보유etf수"].to_numpy(),
    })


def build_raw_full(t_base, t_cmp):
    """현재/기준일 합친 전체 종목 테이블(원 단위 원본 + 정렬용 계산값)."""
    import numpy as np
    m = t_base.merge(t_cmp, on=["티커", "종목명"], how="outer",
                     suffixes=("_현재", "_기준")).fillna(0)
    out = pd.DataFrame({
        "티커": m["티커"].astype(str).str.zfill(6),
        "종목명": m["종목명"],
        "시총_현재": m["시가총액_현재"].astype(float),
        "etf_현재": m["etf보유액_현재"].astype(float),
        "시총_기준": m["시가총액_기준"].astype(float),
        "etf_기준": m["etf보유액_기준"].astype(float),
        "보유ETF수": m["보유etf수_현재"].astype(int),
    })
    sc = out["시총_현재"].to_numpy(); ec = out["etf_현재"].to_numpy()
    sk = out["시총_기준"].to_numpy(); ek = out["etf_기준"].to_numpy()
    out["_편입현재"] = np.where(sc > 0, ec / np.where(sc > 0, sc, 1) * 100, 0.0)
    out["_편입기준"] = np.where(sk > 0, ek / np.where(sk > 0, sk, 1) * 100, 0.0)
    out["_변화"] = out["_편입현재"] - out["_편입기준"]
    out["_보유증감"] = out["etf_현재"] - out["etf_기준"]
    return out


def merge_detail(det_base, det_cmp, names):
    """ETF별 상세(현재/기준일)를 종목-ETF 단위로 합친다."""
    import numpy as np
    EOK = 1e8
    b = det_base.rename(columns={"비중": "비중_현재", "보유액": "보유액_현재", "AUM": "AUM_현재"})
    c = det_cmp.rename(columns={"비중": "비중_기준", "보유액": "보유액_기준", "AUM": "AUM_기준"})
    m = pd.merge(b, c, on=["종목코드", "ETF코드"], how="outer", suffixes=("_b", "_c"))
    m["ETF명"] = m["ETF명_b"].fillna(m["ETF명_c"])
    for col in ["비중_현재", "보유액_현재", "AUM_현재", "비중_기준", "보유액_기준", "AUM_기준"]:
        m[col] = m[col].fillna(0.0)
    cur = m["보유액_현재"].to_numpy(); base = m["보유액_기준"].to_numpy()
    status = np.where((base <= 0) & (cur > 0), "신규",
             np.where((cur <= 0) & (base > 0), "제외(현재없음)",
             np.where(cur > base, "증가", np.where(cur < base, "감소", "유지"))))
    aum_cur = m["AUM_현재"].to_numpy(); aum_base = m["AUM_기준"].to_numpy()
    aum = np.where(aum_cur > 0, aum_cur, aum_base)
    codes = m["종목코드"].astype(str).str.zfill(6)
    out = pd.DataFrame({
        "종목코드": codes,
        "종목명": [names.get(t, t) for t in codes],
        "ETF코드": m["ETF코드"].astype(str).str.zfill(6),
        "ETF명": m["ETF명"].fillna(""),
        "ETF_AUM(억)": (aum / EOK).round(1),
        "비중_현재(%)": m["비중_현재"].round(2),
        "보유액_현재(억)": (m["보유액_현재"] / EOK).round(2),
        "비중_기준일(%)": m["비중_기준"].round(2),
        "보유액_기준일(억)": (m["보유액_기준"] / EOK).round(2),
        "보유액증감(억)": ((m["보유액_현재"] - m["보유액_기준"]) / EOK).round(2),
        "상태": status,
    })
    out = out.sort_values(["종목코드", "보유액_현재(억)"],
                          ascending=[True, False]).reset_index(drop=True)
    return out


# ============================================================
# 엑셀 출력 (값은 엑셀 수식으로 계산 / 2019·2020 호환)
# ============================================================

CALC_COLS = [
    "티커", "종목명",
    "시가총액_현재(원)", "ETF보유액_현재(원)",
    "시가총액_기준일(원)", "ETF보유액_기준일(원)", "보유ETF수",
    "시가총액_현재(억)", "ETF보유액_현재(억)", "ETF보유액_기준일(억)", "보유액증감(억)",
    "편입비중_현재(%)", "편입비중_기준일(%)", "비중변화(%p)",
]


def write_excel(path, base_date, cmp_date, t_base, t_cmp, det_base, det_cmp, names):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    full = build_raw_full(t_base, t_cmp)
    detail = merge_detail(det_base, det_cmp, names)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hfill = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(bold=True, color="FFFFFF")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(1, c)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    WON = ["시가총액_현재(원)", "ETF보유액_현재(원)", "시가총액_기준일(원)",
           "ETF보유액_기준일(원)", "시가총액_현재(억)", "ETF보유액_현재(억)",
           "ETF보유액_기준일(억)", "보유액증감(억)"]
    PCT = ["편입비중_현재(%)", "편입비중_기준일(%)", "비중변화(%p)"]

    def add_calc_sheet(name, df, with_rank):
        ws = wb.create_sheet(name)
        headers = (["순위"] if with_rank else []) + CALC_COLS
        ws.append(headers)
        L = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}
        cC, cD = L["시가총액_현재(원)"], L["ETF보유액_현재(원)"]
        cE, cF = L["시가총액_기준일(원)"], L["ETF보유액_기준일(원)"]
        cCur, cBase = L["편입비중_현재(%)"], L["편입비중_기준일(%)"]
        tk = df["티커"].tolist(); nm = df["종목명"].tolist()
        sc = df["시총_현재"].tolist(); ec = df["etf_현재"].tolist()
        sk = df["시총_기준"].tolist(); ek = df["etf_기준"].tolist()
        cnt = df["보유ETF수"].tolist()
        for idx in range(len(df)):
            i = idx + 2
            vals = (["=ROW()-1"] if with_rank else [])
            vals += [tk[idx], nm[idx], sc[idx], ec[idx], sk[idx], ek[idx], int(cnt[idx])]
            vals += [
                f"={cC}{i}/100000000", f"={cD}{i}/100000000", f"={cF}{i}/100000000",
                f"=({cD}{i}-{cF}{i})/100000000",
                f"=IF({cC}{i}>0,{cD}{i}/{cC}{i}*100,0)",
                f"=IF({cE}{i}>0,{cF}{i}/{cE}{i}*100,0)",
                f"={cCur}{i}-{cBase}{i}",
            ]
            ws.append(vals)
        style_header(ws, len(headers))
        for h in WON:
            for r in range(2, ws.max_row + 1):
                ws[f"{L[h]}{r}"].number_format = "#,##0"
        for h in PCT:
            for r in range(2, ws.max_row + 1):
                ws[f"{L[h]}{r}"].number_format = "0.00"
        for r in range(2, ws.max_row + 1):
            ws[f"{L['티커']}{r}"].number_format = "@"
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 2, 10), 20)
        return ws

    # 0_요약
    ws0 = wb.create_sheet("0_요약")
    for r in [["항목", "값"], ["기준일(현재)", base_date], ["비교 기준일", cmp_date],
              ["분석 종목 수", len(full)],
              ["생성 시각", dt.datetime.now().strftime("%Y-%m-%d %H:%M")],
              ["방법론", "표의 %·억 값은 원본(원) 컬럼에서 엑셀 수식으로 계산됩니다."]]:
        ws0.append(r)
    style_header(ws0, 2)
    ws0.column_dimensions["A"].width = 16
    ws0.column_dimensions["B"].width = 46

    def topN(df, key):
        d = df.sort_values(key, ascending=False).reset_index(drop=True)
        return d.head(TOP_N) if TOP_N else d

    add_calc_sheet("1_시총대비편입비중상위", topN(full, "_편입현재"), True)
    add_calc_sheet("2_편입비중급증상위", topN(full, "_변화"), True)
    add_calc_sheet("3_ETF보유액상위", topN(full, "etf_현재"), True)
    add_calc_sheet("4_ETF보유액급증상위", topN(full, "_보유증감"), True)
    add_calc_sheet("6_전체데이터",
                   full.sort_values("_편입현재", ascending=False).reset_index(drop=True),
                   False)

    Lf = {h: get_column_letter(i + 1) for i, h in enumerate(CALC_COLS)}
    SH = "'6_전체데이터'"

    # ---- 5_조회 (종목 → 요약 한 줄) ----
    ws = wb.create_sheet("5_조회")
    spec = [
        ("종목코드", "티커", "@"), ("종목명", "종목명", None),
        ("현재 편입비중(%)", "편입비중_현재(%)", "0.00"),
        ("기준일 편입비중(%)", "편입비중_기준일(%)", "0.00"),
        ("비중변화(%p)", "비중변화(%p)", "0.00"),
        ("ETF보유액_현재(억)", "ETF보유액_현재(억)", "#,##0"),
        ("ETF보유액_기준일(억)", "ETF보유액_기준일(억)", "#,##0"),
        ("보유액증감(억)", "보유액증감(억)", "#,##0"),
        ("시가총액(억)", "시가총액_현재(억)", "#,##0"),
        ("보유ETF수", "보유ETF수", "0"),
    ]
    hdr = ["입력(코드/종목명)"] + [s[0] for s in spec]
    ws.append(hdr)
    ws.append(["← 노란 칸에 종목코드(예 005930) 나 종목명을 한 줄에 하나씩 입력"])
    style_header(ws, len(hdr))
    yellow = PatternFill("solid", fgColor="FFF2CC")
    helper_col = get_column_letter(len(hdr) + 2)
    for r in range(3, 43):
        ws.cell(r, 1).fill = yellow
        ws.cell(r, 1).number_format = "@"
        ws[f"{helper_col}{r}"] = (
            f'=IFERROR(MATCH($A{r},{SH}!${Lf["종목명"]}:${Lf["종목명"]},0),'
            f'IFERROR(MATCH($A{r},{SH}!${Lf["티커"]}:${Lf["티커"]},0),'
            f'IFERROR(MATCH(TEXT($A{r},"000000"),{SH}!${Lf["티커"]}:${Lf["티커"]},0),NA())))')
        for j, (_, srccol, numfmt) in enumerate(spec):
            col = get_column_letter(j + 2)
            src = Lf[srccol]
            fb = '"못 찾음"' if srccol == "티커" else '""'
            ws[f"{col}{r}"] = (f'=IF($A{r}="","",'
                               f'IFERROR(INDEX({SH}!${src}:${src},${helper_col}{r}),{fb}))')
            if numfmt:
                ws[f"{col}{r}"].number_format = numfmt
    ws.column_dimensions[helper_col].hidden = True
    ws.column_dimensions["A"].width = 18
    for j in range(len(spec)):
        ws.column_dimensions[get_column_letter(j + 2)].width = 16

    # ---- 8_ETF상세 (종목별ETF조회의 데이터 원본) ----
    DCOLS = ["종목코드", "종목명", "ETF코드", "ETF명", "ETF_AUM(억)",
             "비중_현재(%)", "보유액_현재(억)", "비중_기준일(%)", "보유액_기준일(억)",
             "보유액증감(억)", "상태"]
    wsd = wb.create_sheet("8_ETF상세")
    wsd.append(DCOLS)
    arrs = {c: detail[c].tolist() for c in DCOLS}
    for idx in range(len(detail)):
        wsd.append([arrs[c][idx] for c in DCOLS])
    style_header(wsd, len(DCOLS))
    DL = {h: get_column_letter(i + 1) for i, h in enumerate(DCOLS)}
    for r in range(2, wsd.max_row + 1):
        wsd[f"{DL['종목코드']}{r}"].number_format = "@"
        wsd[f"{DL['ETF코드']}{r}"].number_format = "@"
        wsd[f"{DL['ETF_AUM(억)']}{r}"].number_format = "#,##0"
        for h in ["비중_현재(%)", "보유액_현재(억)", "비중_기준일(%)",
                  "보유액_기준일(억)", "보유액증감(억)"]:
            wsd[f"{DL[h]}{r}"].number_format = "0.00"
    for i, h in enumerate(DCOLS, 1):
        wsd.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 2, 10), 22)

    # ---- 7_종목별ETF조회 (종목 → 그 종목 보유 ETF 전체 나열) ----
    D = "'8_ETF상세'"
    ws7 = wb.create_sheet("7_종목별ETF조회")
    ws7["A1"] = "종목코드/종목명 →"
    ws7["A1"].font = Font(bold=True)
    ws7["B1"].fill = yellow
    ws7["B1"].number_format = "@"
    # 헬퍼: L1 해석된 코드 / M1 첫 행 / N1 개수
    ws7["L1"] = (f'=IFERROR(INDEX({SH}!$A:$A,MATCH($B$1,{SH}!$B:$B,0)),'
                 f'IFERROR(INDEX({SH}!$A:$A,MATCH($B$1,{SH}!$A:$A,0)),'
                 f'IFERROR(INDEX({SH}!$A:$A,MATCH(TEXT($B$1,"000000"),{SH}!$A:$A,0)),"")))')
    ws7["M1"] = f'=IFERROR(MATCH($L$1,{D}!${DL["종목코드"]}:${DL["종목코드"]},0),0)'
    ws7["N1"] = f'=IF($L$1="",0,COUNTIF({D}!${DL["종목코드"]}:${DL["종목코드"]},$L$1))'
    ws7["D1"] = '=IF($L$1="","종목을 입력하세요", "보유 ETF "&$N$1&"개")'
    out_spec = [
        ("ETF코드", DL["ETF코드"], "@"), ("ETF명", DL["ETF명"], None),
        ("ETF AUM(억)", DL["ETF_AUM(억)"], "#,##0"),
        ("현재 비중(%)", DL["비중_현재(%)"], "0.00"),
        ("현재 보유액(억)", DL["보유액_현재(억)"], "0.00"),
        ("기준일 비중(%)", DL["비중_기준일(%)"], "0.00"),
        ("기준일 보유액(억)", DL["보유액_기준일(억)"], "0.00"),
        ("보유액증감(억)", DL["보유액증감(억)"], "0.00"),
        ("상태", DL["상태"], None),
    ]
    hrow = 3
    ws7.append([])  # row2 빈 줄
    ws7.cell(hrow, 1, "순위")
    for j, (label, _, _) in enumerate(out_spec):
        ws7.cell(hrow, j + 2, label)
    for c in range(1, len(out_spec) + 2):
        cell = ws7.cell(hrow, c)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws7.freeze_panes = f"A{hrow + 1}"
    NSHOW = 320
    for r in range(hrow + 1, hrow + 1 + NSHOW):
        k = r - hrow  # 1부터
        ws7.cell(r, 1).value = (f'=IF(AND($L$1<>"",{k}<=$N$1),{k},"")')
        for j, (_, sc_, numfmt) in enumerate(out_spec):
            col = get_column_letter(j + 2)
            ws7[f"{col}{r}"] = (f'=IF(AND($L$1<>"",{k}<=$N$1),'
                                f'INDEX({D}!${sc_}:${sc_},$M$1+{k}-1),"")')
            if numfmt:
                ws7[f"{col}{r}"].number_format = numfmt
    for cc in ["L", "M", "N"]:
        ws7.column_dimensions[cc].hidden = True
    ws7.column_dimensions["A"].width = 6
    ws7.column_dimensions["B"].width = 12
    ws7.column_dimensions["C"].width = 30
    for j in range(2, len(out_spec) + 1):
        ws7.column_dimensions[get_column_letter(j + 2)].width = 15

    order = ["0_요약", "1_시총대비편입비중상위", "2_편입비중급증상위",
             "3_ETF보유액상위", "4_ETF보유액급증상위", "5_조회",
             "7_종목별ETF조회", "6_전체데이터", "8_ETF상세"]
    wb._sheets.sort(key=lambda x: order.index(x.title) if x.title in order else 99)
    wb.save(path)


def update_web_data(date, agg, mcap, detail, names):
    """웹 뷰어용 데이터 출력.
    - data.json : 날짜 목록만(경량)
    - stocks_<date>.json : 그 날짜의 '종목 집계'(가벼움) — 날짜별 분리
    - detail_<date>.json : 그 날짜의 'ETF별 상세'(무거움) — 날짜별 분리 저장
    금액 단위는 억원."""
    EOK = 1e8
    mc = mcap["시가총액"]
    stocks = {}
    for code in agg.index:
        if code not in mc.index:
            continue
        rec = {
            "name": names.get(code, code),
            "mcap": round(float(mc.loc[code]) / EOK),
            "val": round(float(agg.loc[code, "etf보유액"]) / EOK),
            "n": int(agg.loc[code, "보유etf수"]),
        }
        if "종가" in mcap.columns:
            rec["price"] = int(round(float(mcap.loc[code, "종가"])))
        if "상장주식수" in mcap.columns:
            rec["shares"] = int(float(mcap.loc[code, "상장주식수"]))
        stocks[code] = rec
    # 종목 집계: 날짜별 파일(stocks_<date>.json)로 분리 — data.json 비대화 방지
    with open(os.path.join(WEB_DIR, f"stocks_{date}.json"), "w", encoding="utf-8") as f:
        json.dump(stocks, f, ensure_ascii=False)
    # data.json: 날짜 목록만(경량). 옛 형식의 stocks/detail 키는 제거.
    djson = os.path.join(WEB_DIR, "data.json")
    data = {"dates": []}
    if os.path.exists(djson):
        try:
            data = json.load(open(djson, encoding="utf-8"))
        except Exception:
            data = {"dates": []}
    data.pop("detail", None)
    data.pop("stocks", None)
    data["dates"] = sorted(set(data.get("dates", [])) | {date})
    data["generated"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M") + " 갱신"
    with open(djson, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)

    # detail_<date>.json
    codes = detail["종목코드"].tolist()
    etfs = detail["ETF코드"].tolist()
    enames = detail["ETF명"].tolist()
    ws = detail["비중"].tolist()
    vals = detail["보유액"].tolist()
    aums = detail["AUM"].tolist()
    qs = detail["계약수"].tolist() if "계약수" in detail.columns else [0]*len(detail)
    # 압축형: ETF 이름/AUM은 "_etf"에 한 번만, 각 행은 e/w(비중)/v(보유액)/q(계약수)
    det = {"_etf": {}}
    for i in range(len(detail)):
        e = etfs[i]
        if e not in det["_etf"]:
            det["_etf"][e] = {"name": enames[i], "aum": round(float(aums[i]) / EOK)}
        det.setdefault(codes[i], []).append({
            "e": e,
            "w": round(float(ws[i]), 2),
            "v": round(float(vals[i]) / EOK, 1),
            "q": round(float(qs[i]), 2),
        })
    for c in det:
        if c == "_etf":
            continue
        det[c].sort(key=lambda x: -x["v"])
    if INCLUDE_ETF:
        det["_withetf"] = 1   # ETF 보유분까지 받았다는 표시(_is_done 재크롤 판정용)
    with open(os.path.join(WEB_DIR, f"detail_{date}.json"), "w", encoding="utf-8") as f:
        json.dump(det, f, ensure_ascii=False)


# ============================================================
# 메인
# ============================================================

def _make_excel(base, store):
    """최근 크롤 날짜(base) vs COMPARE_DATE 로 엑셀 1개 생성."""
    cmp = COMPARE_DATE
    if base == cmp:
        print("  엑셀 생략(기준일=비교일)")
        return

    def get(date):
        if date in store:
            return store[date]
        mc = get_market_cap(date)
        ag, de = aggregate_for_date(date, set(mc.index))
        nm = build_name_map(set(ag.index))
        return (ag, mc, de, nm)

    ab, mb, db, nb = get(base)
    ac, mc, dc, nc = get(cmp)
    names = {**nc, **nb}
    t_base = make_ratio_table(ab, mb, names)
    t_cmp = make_ratio_table(ac, mc, names)
    out = os.path.join(OUTPUT_DIR, f"ETF편입분석_{base}_vs_{cmp}.xlsx")
    write_excel(out, base, cmp, t_base, t_cmp, db, dc, names)
    print(f"  엑셀: {out}")


def _is_done(date):
    """'새 스키마'로 받아졌는지 — stocks 파일에 price(종가)가 있으면 완료.
    옛 스키마(price 없음)는 미완료로 봐서 자동 재크롤한다."""
    p = os.path.join(WEB_DIR, f"stocks_{date}.json")
    if not os.path.exists(p):
        return False
    try:
        d = json.load(open(p, encoding="utf-8"))
        has_price = False
        for v in d.values():
            has_price = "price" in v
            break
        if not has_price:
            return False
        if INCLUDE_ETF:
            dp = os.path.join(WEB_DIR, f"detail_{date}.json")
            if not os.path.exists(dp):
                return False
            return bool(json.load(open(dp, encoding="utf-8")).get("_withetf"))
        return True
    except Exception:
        return False


def trading_days_of_year(year):
    """해당 연도의 실제 거래일 목록(YYYYMMDD). 최근 영업일까지만."""
    y = int(year)
    end = resolve_base_date()
    if int(end[:4]) > y:
        end = f"{y}1231"
    try:
        days = stock.get_previous_business_days(fromdate=f"{y}0101", todate=end)
    except Exception as e:
        print("  거래일 목록 조회 실패:", e)
        return []
    out = []
    for d in days:
        try:
            out.append(d.strftime("%Y%m%d"))
        except Exception:
            out.append(str(d).replace("-", "")[:8])
    # 오늘 날짜는 제외(장중/당일분은 --year 배치에서 받지 않음)
    _today = dt.date.today().strftime("%Y%m%d")
    out = [d for d in out if d < _today]
    return sorted(set(out))


def _merge_etf_codes(codes):
    """상장 ETF 전체 코드를 data/etf_codes.json에 누적 저장(웹 시장전체 탭 필터용).
    보유내용과 무관한 '진짜 ETF 명단'이라 채권형 ETF도 확실히 걸러진다."""
    p = os.path.join(WEB_DIR, "etf_codes.json")
    cur = set()
    if os.path.exists(p):
        try:
            cur = set(json.load(open(p, encoding="utf-8")))
        except Exception:
            cur = set()
    new = cur | {str(c) for c in codes if c}
    if new != cur:
        try:
            with open(p, "w", encoding="utf-8") as f:
                json.dump(sorted(new), f, ensure_ascii=False)
        except Exception:
            pass


def _crawl_one(date):
    """한 날짜 수집+웹저장. 성공 True / throttle·실패 False."""
    mcap = get_market_cap(date)
    if mcap.empty:
        print(f"  WARN {date}: 시총 빈 응답(throttle/휴장). 이 날짜 보류.")
        return False
    etf_univ = None
    if INCLUDE_ETF:   # ETF를 '종목'처럼 시총 유니버스에 합쳐 ETF-in-ETF 보유도 잡음
        etf_univ = get_etf_universe(date)
        if etf_univ:
            _merge_etf_codes(etf_univ.keys())   # 전체 ETF 명단 갱신
            add = {c: v for c, v in etf_univ.items()
                   if c not in mcap.index and v["시가총액"] > 0}
            if add:
                edf = pd.DataFrame.from_dict(
                    {c: {"종가": v["종가"], "시가총액": v["시가총액"],
                         "상장주식수": v["상장주식수"]} for c, v in add.items()},
                    orient="index")
                mcap = pd.concat([mcap, edf])
    try:
        agg, detail = aggregate_for_date(date, set(mcap.index))
    except ThrottledError:
        return False
    known = None
    if etf_univ:   # ETF 보유분 이름은 미리 알고 있으니 조회 생략
        known = {c: etf_univ[c]["name"] for c in etf_univ
                 if isinstance(etf_univ[c].get("name"), str) and etf_univ[c]["name"]}
    names = build_name_map(set(agg.index), known=known)
    update_web_data(date, agg, mcap, detail, names)
    print(f"  웹 데이터 갱신 완료 ({date})")
    return True


def run_year_batch():
    """--year 모드: 거래일 전체 중 '안 받은 것'만, 한 번에 BATCH_PER_RUN개씩.
    종료코드 0=전부완료 / 2=배치끝(더남음) / 3=throttle(길게 쉬고 재시도)."""
    all_dates = trading_days_of_year(YEAR)
    if not all_dates:
        print("거래일 목록을 못 받았습니다(throttle 가능). 잠시 후 재시도.")
        return 3
    todo = [d for d in all_dates if not _is_done(d)]
    done = len(all_dates) - len(todo)
    print(f"\n[연도 {YEAR}] 거래일 {len(all_dates)}개 / 완료 {done} / 남음 {len(todo)}")
    if not todo:
        print("=== 모든 거래일 완료! ===")
        return 0
    batch = list(reversed(todo))[:BATCH_PER_RUN]   # 최신 날짜부터(역순)
    print(f"이번 실행 배치({len(batch)}개): {batch}\n")
    throttled = False
    for n, date in enumerate(batch, 1):
        print(f"===== [{n}/{len(batch)}] {date} =====")
        if not _crawl_one(date):
            throttled = True
            print("  -> throttle로 판단, 이번 배치 중단(잠시 쉬었다 재시도).")
            break
        if REST_BETWEEN and n < len(batch):
            print(f"  ({REST_BETWEEN}초 쉬는 중... throttle 방지)")
            time.sleep(REST_BETWEEN)
    remaining = [d for d in all_dates if not _is_done(d)]
    print(f"\n남은 거래일: {len(remaining)}개")
    if not remaining:
        print("=== 모든 거래일 완료! ===")
        return 0
    return 3 if throttled else 2


def main():
    if not ensure_krx_login():
        return 1

    if YEAR:
        return run_year_batch()

    dates = [resolve_base_date()] if AUTO else ask_dates_multi()
    dates = sorted(set(dates))
    print(f"\n크롤 대상 날짜 ({len(dates)}개): {dates}\n")

    store = {}
    for n, date in enumerate(dates, 1):
        print(f"===== [{n}/{len(dates)}] {date} =====")
        if _crawl_one(date):
            store[date] = True
        if REST_BETWEEN and n < len(dates):
            print(f"  ({REST_BETWEEN}초 쉬는 중... throttle 방지)")
            time.sleep(REST_BETWEEN)

    print("\n[웹] data.json / detail_*.json 모두 갱신 완료.")
    print("\n완료!")
    return 0


if __name__ == "__main__":
    _code = 1
    try:
        _code = main() or 0
    except ModuleNotFoundError as e:
        print("\n[오류] 필요한 패키지가 없습니다:", e)
        print("    pip install pykrx openpyxl pandas tqdm")
    except Exception:
        import traceback
        print("\n[오류] 실행 중 문제가 발생했습니다. 아래 내용을 캡쳐해 알려주세요:\n")
        traceback.print_exc()
    if not AUTO and not YEAR:
        input("\n끝났습니다. 이 창을 닫으려면 Enter 키를 누르세요...")
    sys.exit(_code)
